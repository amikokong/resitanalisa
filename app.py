"""
app.py  — Resit Analisa (Streamlit + Gemini + Google Drive OAuth)
-----------------------------------------------------------------
Aliran kerja automatik:
  - Log masuk Google sekali (OAuth)
  - App BACA fail induk terus dari Drive anda
  - Upload resit baru -> AI ekstrak -> anti-pendua
  - App TULIS fail induk terus ke Drive (tiada download/upload manual)

Skop Drive: drive.file -> app HANYA nampak fail yang ia sendiri cipta.
            (Tidak boleh intai fail lain dalam Drive anda.)

Jalankan:
    pip install -r requirements.txt
    streamlit run app.py

Secrets (Streamlit Cloud: Settings > Secrets):
    GEMINI_API_KEY      = "AIza..."
    GOOGLE_CLIENT_ID    = "xxxx.apps.googleusercontent.com"
    GOOGLE_CLIENT_SECRET= "GOCSPX-xxxx"
    REDIRECT_URI        = "https://resitanalisa.streamlit.app"   # URL app anda

Dibangunkan oleh: Sulaiman Osman  (sulaimanosman03@gmail.com)
"""

import json
import os
import time

os.environ["OAUTHLIB_RELAX_TOKEN_SCOPE"] = "1"  # elak ralat 'scope changed' bila tambah skop email
from io import BytesIO

import streamlit as st
import pandas as pd
from PIL import Image
from google import genai
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

MODEL_NAME = "gemini-2.5-flash"
MONEY_FORMAT = "#,##0.00"
BOLD = Font(bold=True)

AUTHOR_NAME = "Sulaiman Osman"
AUTHOR_EMAIL = "sulaimanosman03@gmail.com"
AUTHOR_YEAR = "2026"

APP_NAME = "AI Expenditure Tracking System"
APP_SHORT = "ETS"

TOS_URL = "https://docs.google.com/document/d/e/2PACX-1vRrOFJo1rzqQEPnANY6QtHNoOb2RYfwiFSJWlIszCEixup6sQV2rMHTBLxCpJyHI8XcYTdeB46XLu7T/pub"
PRIVACY_URL = "https://docs.google.com/document/d/e/2PACX-1vRzxxLKrCCoHE1PKGnr0W_XZ9aFwA5BGO-n4CweajSmxDsrfxRii7P6N6U8GP0bOrS4TZVjZno8qfdn/pub"

MASTER_PREFIX = "fail_induk_perakaunan"
AUTOBACKUP_NAME = "fail_induk_perakaunan_AUTOBACKUP.xlsx"
PROFILE_FILENAME = "ets_profile.json"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SCOPES = [
    "https://www.googleapis.com/auth/drive.file",
    "openid",
    "https://www.googleapis.com/auth/userinfo.email",
]

# Nama bulan Bahasa Melayu (untuk folder bulanan)
BULAN_MY = ["Januari", "Februari", "Mac", "April", "Mei", "Jun",
            "Julai", "Ogos", "September", "Oktober", "November", "Disember"]

FIELDS = [
    "transaction_date", "vendor_name", "receipt_no", "description",
    "amount", "debit_account", "credit_account", "currency",
]


# ==================================================
# HELPER EXCEL
# ==================================================
def write_headers(ws, headers, row=1):
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center")


def apply_money_format(ws, columns, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.column_letter in columns and isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FORMAT


def autofit(wb, width=22):
    for ws in wb.worksheets:
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = width


def add_footer(ws):
    last = ws.max_row + 2
    ws.cell(row=last, column=1, value=f"\u00A9 {AUTHOR_YEAR} Dibangunkan oleh: {AUTHOR_NAME}")
    link = ws.cell(row=last + 1, column=1, value=AUTHOR_EMAIL)
    link.hyperlink = f"mailto:{AUTHOR_EMAIL}"
    link.font = Font(color="0563C1", underline="single")


# ==================================================
# ANTI-PENDUA
# ==================================================
def _vendor(t):
    return str(t.get("vendor_name", "")).strip().lower()


def classify(txn, existing):
    """new = tambah · exact = pendua tepat (langkau) · maybe = sahkan manusia."""
    receipt_no = str(txn.get("receipt_no", "")).strip()
    if receipt_no:  # Lapisan 1: nombor resit
        for e in existing:
            if _vendor(e) == _vendor(txn) and str(e.get("receipt_no", "")).strip() == receipt_no:
                return "exact"
        return "new"
    # Lapisan 2: tiada nombor resit
    amt = round(float(txn.get("amount", 0) or 0), 2)
    date = str(txn.get("transaction_date", "")).strip()
    for e in existing:
        if (str(e.get("transaction_date", "")).strip() == date
                and _vendor(e) == _vendor(txn)
                and round(float(e.get("amount", 0) or 0), 2) == amt):
            return "maybe"
    return "new"


# ==================================================
# AI
# ==================================================
def extract_receipt_data(image, max_retries=4):
    client = genai.Client(api_key=st.secrets["GEMINI_API_KEY"])
    prompt = """
    Baca resit ini dan pulangkan HANYA objek JSON (tiada teks lain, tiada markdown).
    Cari nombor resit / invois / bil untuk 'receipt_no'. Jika tiada, letak "".
    Format:
    {
      "transaction_date": "YYYY-MM-DD",
      "vendor_name": "...",
      "receipt_no": "...",
      "description": "...",
      "amount": 0.00,
      "debit_account": "...",
      "credit_account": "Cash",
      "currency": "MYR"
    }
    """
    last_err = None
    for attempt in range(max_retries):
        try:
            response = client.models.generate_content(
                model=MODEL_NAME,
                contents=[prompt, image],
                config={"response_mime_type": "application/json"},
            )
            return json.loads(response.text)
        except Exception as e:
            last_err = e
            msg = str(e).lower()
            # Cuba semula HANYA untuk ralat sementara (server sibuk)
            if any(k in msg for k in ["503", "unavailable", "overload", "429", "rate"]):
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)  # tunggu 1, 2, 4 saat (backoff)
                    continue
            raise  # ralat lain (cth JSON rosak) -> terus naikkan
    raise last_err


# ==================================================
# EXCEL: baca & bina
# ==================================================
def read_master_bytes(buf):
    wb = load_workbook(buf)
    if "Transactions" not in wb.sheetnames:
        return []
    ws = wb["Transactions"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        out.append({FIELDS[i]: (row[i] if i < len(row) else "") for i in range(len(FIELDS))})
    return out


def build_master_excel(transactions, profile=None):
    profile = profile or {}
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Transactions"
    write_headers(ws1, ["Date", "Vendor", "Receipt No", "Description", "Amount",
                        "Debit Account", "Credit Account", "Currency"])
    for t in transactions:
        ws1.append([
            t.get("transaction_date", ""), t.get("vendor_name", ""),
            t.get("receipt_no", ""), t.get("description", ""),
            float(t.get("amount", 0) or 0), t.get("debit_account", ""),
            t.get("credit_account", ""), t.get("currency", "MYR"),
        ])
    apply_money_format(ws1, ["E"])

    ws2 = wb.create_sheet("Journal Entry")
    write_headers(ws2, ["Date", "Description", "Account", "Debit", "Credit"])
    for t in transactions:
        amt = float(t.get("amount", 0) or 0)
        ws2.append([t.get("transaction_date", ""), t.get("description", ""),
                    t.get("debit_account", ""), amt, None])
        ws2.append([t.get("transaction_date", ""), t.get("description", ""),
                    t.get("credit_account", ""), None, amt])
    apply_money_format(ws2, ["D", "E"])

    df = pd.DataFrame(transactions)
    if not df.empty:
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0)
        df["transaction_date"] = pd.to_datetime(df["transaction_date"], errors="coerce")
        df["month"] = df["transaction_date"].dt.strftime("%Y-%m")

    # ---- Monthly Summary (dengan header profil) ----
    ws3 = wb.create_sheet("Monthly Summary")
    ws3["A1"] = f"{APP_NAME} ({APP_SHORT})"
    ws3["A1"].font = Font(bold=True, size=14, color="14305C")
    prow = 2
    if profile.get("full_name"):
        ws3.cell(row=prow, column=1, value=f"Nama: {profile['full_name']}").font = BOLD
        prow += 1
    if profile.get("company"):
        ws3.cell(row=prow, column=1, value=f"Syarikat: {profile['company']}").font = BOLD
        prow += 1
    ws3.cell(row=prow, column=1,
             value=f"Laporan dijana: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}")
    hdr = prow + 2
    write_headers(ws3, ["Month", "Total Expense (RM)", "Bil. Transaksi"], row=hdr)
    if not df.empty:
        monthly = df.groupby("month").agg(total=("amount", "sum"),
                                          count=("amount", "count")).reset_index()
        for _, r in monthly.iterrows():
            ws3.append([r["month"], float(r["total"]), int(r["count"])])
    apply_money_format(ws3, ["B"], start_row=hdr + 1)
    add_footer(ws3)

    ws4 = wb.create_sheet("By Account")
    write_headers(ws4, ["Debit Account", "Total (RM)"])
    if not df.empty:
        by_acc = df.groupby("debit_account")["amount"].sum().reset_index()
        for _, r in by_acc.iterrows():
            ws4.append([r["debit_account"], float(r["amount"])])
    apply_money_format(ws4, ["B"])

    autofit(wb)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ==================================================
# GOOGLE DRIVE (OAuth)
# ==================================================
def get_flow():
    return Flow.from_client_config(
        {
            "web": {
                "client_id": st.secrets["GOOGLE_CLIENT_ID"],
                "client_secret": st.secrets["GOOGLE_CLIENT_SECRET"],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [st.secrets["REDIRECT_URI"]],
            }
        },
        scopes=SCOPES,
        redirect_uri=st.secrets["REDIRECT_URI"],
        autogenerate_code_verifier=False,  # matikan PKCE (elak ralat "Missing code verifier")
    )


def drive_service(creds):
    return build("drive", "v3", credentials=creds)


def get_user_email(creds):
    """Dapatkan emel pengguna yang log masuk (untuk kawalan akses/trial)."""
    try:
        oauth2 = build("oauth2", "v2", credentials=creds)
        return oauth2.userinfo().get().execute().get("email", "").lower()
    except Exception:
        return ""


def check_access(email):
    """
    Semak akses pengguna ikut senarai dalam Secrets [access].
    Pulang: (status, hari_tinggal)
      status: 'full' | 'trial' | 'expired' | 'denied'
    Format Secrets:
      [access]
      "user@gmail.com" = "paid"                 -> akses penuh
      "user@gmail.com" = "trial:2026-06-09:3"   -> trial mula 9 Jun, 3 hari
    """
    try:
        access = dict(st.secrets.get("access", {}))
    except Exception:
        access = {}
    rule = access.get(email)
    if not rule:
        return ("denied", None)
    rule = str(rule).strip().lower()
    if rule == "paid":
        return ("full", None)
    if rule.startswith("trial:"):
        try:
            _, start_s, days_s = rule.split(":")
            start = pd.Timestamp(start_s).normalize()
            expiry = start + pd.Timedelta(days=int(days_s))
            today = pd.Timestamp.now().normalize()
            if today <= expiry:
                return ("trial", (expiry - today).days)
            return ("expired", None)
        except Exception:
            return ("denied", None)
    return ("denied", None)


def _list_in_folder(service, folder_id):
    """Senarai semua fail dalam folder (untuk tapis dalam Python)."""
    res = service.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        spaces="drive", fields="files(id,name)", pageSize=1000,
    ).execute()
    return res.get("files", [])


def find_latest_master(service):
    """Cari fail induk TERKINI (ikut masa ubah suai). Tapis dalam Python
    untuk elak isu 'name contains' Drive dengan nama berunderscore."""
    res = service.files().list(
        q=("mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
           "and trashed=false"),
        spaces="drive", orderBy="modifiedTime desc",
        fields="files(id,name,modifiedTime)", pageSize=100,
    ).execute()
    for f in res.get("files", []):
        if f["name"].startswith("fail_induk_perakaunan"):
            return f
    return None


def download_master(service, file_id):
    req = service.files().get_media(fileId=file_id)
    buf = BytesIO()
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf


def monthly_folder_name():
    """Nama folder bulanan, cth: 'Resit bagi bulan Jun 2026'."""
    now = pd.Timestamp.now()
    return f"Resit bagi bulan {BULAN_MY[now.month - 1]} {now.year}"


def get_or_create_folder(service, name):
    """Cari folder ikut nama; cipta jika belum wujud. Pulang ID folder."""
    q = (f"name='{name}' and mimeType='application/vnd.google-apps.folder' "
         f"and trashed=false")
    res = service.files().list(q=q, spaces="drive", fields="files(id,name)").execute()
    files = res.get("files", [])
    if files:
        return files[0]["id"]
    meta = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
    folder = service.files().create(body=meta, fields="id").execute()
    return folder["id"]


def next_filename(service, folder_id):
    """Jana nama fail bertarikh + nombor, cth: fail_induk_perakaunan_08062026(01).xlsx"""
    datestr = pd.Timestamp.now().strftime("%d%m%Y")
    base = f"fail_induk_perakaunan_{datestr}"
    count = sum(1 for f in _list_in_folder(service, folder_id) if f["name"].startswith(base))
    return f"{base}({count + 1:02d}).xlsx"


def save_master_in_folder(service, excel_bytes, folder_id, filename):
    """Cipta fail Excel BARU dalam folder bulanan (setiap simpan = snapshot)."""
    media = MediaIoBaseUpload(BytesIO(excel_bytes), mimetype=XLSX_MIME, resumable=True)
    meta = {"name": filename, "mimeType": XLSX_MIME, "parents": [folder_id]}
    created = service.files().create(body=meta, media_body=media, fields="id,name").execute()
    return created


def auto_backup(service, excel_bytes):
    """Backup BERGULIR: kemas kini SATU fail backup (tak cipta banyak salinan)."""
    folder_id = get_or_create_folder(service, monthly_folder_name())
    q = f"name='{AUTOBACKUP_NAME}' and '{folder_id}' in parents and trashed=false"
    res = service.files().list(q=q, spaces="drive", fields="files(id)").execute()
    files = res.get("files", [])
    media = MediaIoBaseUpload(BytesIO(excel_bytes), mimetype=XLSX_MIME, resumable=True)
    if files:
        service.files().update(fileId=files[0]["id"], media_body=media).execute()
    else:
        meta = {"name": AUTOBACKUP_NAME, "mimeType": XLSX_MIME, "parents": [folder_id]}
        service.files().create(body=meta, media_body=media, fields="id").execute()


def maybe_autobackup(service):
    """Jalankan auto-backup jika diaktifkan. Senyap jika gagal (tak ganggu aliran)."""
    if st.session_state.get("auto_backup") and st.session_state.get("transactions"):
        try:
            auto_backup(service, build_master_excel(
                st.session_state.transactions, st.session_state.get("profile")).getvalue())
        except Exception:
            pass


# ---- Profil pengguna (disimpan sebagai JSON di Drive) ----
def save_profile(service, profile):
    data = json.dumps(profile).encode("utf-8")
    media = MediaIoBaseUpload(BytesIO(data), mimetype="application/json", resumable=True)
    q = f"name='{PROFILE_FILENAME}' and trashed=false"
    res = service.files().list(q=q, spaces="drive", fields="files(id)").execute()
    files = res.get("files", [])
    if files:
        service.files().update(fileId=files[0]["id"], media_body=media).execute()
    else:
        meta = {"name": PROFILE_FILENAME, "mimeType": "application/json"}
        service.files().create(body=meta, media_body=media, fields="id").execute()


def load_profile(service):
    try:
        q = f"name='{PROFILE_FILENAME}' and trashed=false"
        res = service.files().list(q=q, spaces="drive", fields="files(id)").execute()
        files = res.get("files", [])
        if not files:
            return {}
        req = service.files().get_media(fileId=files[0]["id"])
        buf = BytesIO()
        dl = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
        buf.seek(0)
        return json.loads(buf.read().decode("utf-8"))
    except Exception:
        return {}


# ---- Arkib dokumen sumber (folder "Resit Softcopy YYYY") ----
def softcopy_folder_name():
    """Folder arkib tahunan untuk dokumen sumber, cth: 'Resit Softcopy 2026'."""
    return f"Resit Softcopy {pd.Timestamp.now().year}"


def count_archive_today(service, folder_id):
    """Bilangan fail arkib untuk tarikh hari ini (untuk nombor turutan)."""
    datestr = pd.Timestamp.now().strftime("%d%m%Y")
    base = f"resit_{datestr}"
    return sum(1 for f in _list_in_folder(service, folder_id) if f["name"].startswith(base))


def save_archive_file(service, data_bytes, folder_id, seq, ext, mimetype):
    """Simpan satu dokumen sumber: resit_DDMMYYYY(seq).ext"""
    datestr = pd.Timestamp.now().strftime("%d%m%Y")
    filename = f"resit_{datestr}({seq}).{ext}"
    media = MediaIoBaseUpload(BytesIO(data_bytes), mimetype=mimetype, resumable=True)
    meta = {"name": filename, "parents": [folder_id]}
    service.files().create(body=meta, media_body=media, fields="id").execute()
    return filename


def make_voucher_image(txn):
    """Jana imej baucar PNG untuk transaksi manual (tiada resit fizikal)."""
    from PIL import ImageDraw, ImageFont
    W, H = 620, 420
    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)
    try:
        fb = ImageFont.truetype("DejaVuSans-Bold.ttf", 22)
        ff = ImageFont.truetype("DejaVuSans.ttf", 17)
    except Exception:
        fb = ImageFont.load_default()
        ff = ImageFont.load_default()
    d.rectangle([0, 0, W, 62], fill=(20, 48, 92))
    d.text((22, 19), "BAUCAR TRANSAKSI MANUAL", fill="white", font=fb)
    rows = [
        ("No. Siri", txn.get("receipt_no", "")),
        ("Tarikh", txn.get("transaction_date", "")),
        ("Vendor / Tempat", txn.get("vendor_name", "")),
        ("Keterangan", txn.get("description", "")),
        ("Jumlah (RM)", f"{float(txn.get('amount', 0) or 0):.2f}"),
        ("Akaun Debit", txn.get("debit_account", "")),
        ("Akaun Kredit", txn.get("credit_account", "")),
    ]
    y = 88
    for label, val in rows:
        d.text((22, y), f"{label}:", fill=(91, 107, 133), font=ff)
        d.text((210, y), str(val)[:48], fill=(27, 38, 56), font=ff)
        y += 40
    d.text((22, H - 32), "Transaksi tanpa resit fizikal — direkod secara manual",
           fill=(154, 166, 188), font=ff)
    buf = BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


# ==================================================
# ANTARA MUKA STREAMLIT
# ==================================================
st.set_page_config(page_title="ETS — AI Expenditure Tracking", page_icon="🧾", layout="centered")

# ==================================================
# GAYA / TEMA (CSS tersuai — rupa fintech profesional)
# ==================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@500;600;700&family=IBM+Plex+Sans:wght@400;500;600&display=swap');

:root{
  --navy:#14305C; --navy2:#1E4A86; --accent:#2F8FBF;
  --ink:#1B2638; --muted:#5B6B85; --line:#E4E9F2;
  --bg:#EEF2F9; --card:#FFFFFF;
}
.stApp{ background:
  radial-gradient(900px 500px at 100% -10%, #E3ECFa 0%, transparent 55%),
  linear-gradient(180deg,#EEF2F9 0%,#F6F8FC 100%); }
html, body, [class*="css"]{ font-family:'IBM Plex Sans',sans-serif; color:var(--ink); }
h1,h2,h3,h4{ font-family:'Sora',sans-serif !important; color:var(--navy) !important; letter-spacing:-.01em; }

#MainMenu, footer, [data-testid="stHeader"]{ visibility:hidden; }
.block-container{ padding-top:1.4rem; max-width:780px; }

/* Hero */
.hero{ background:linear-gradient(120deg,var(--navy) 0%,var(--navy2) 70%,var(--accent) 130%);
  border-radius:20px; padding:26px 30px; color:#fff; display:flex; align-items:center; gap:18px;
  box-shadow:0 14px 34px rgba(20,48,92,.28); margin-bottom:22px; }
.hero .mark{ font-size:34px; background:rgba(255,255,255,.14); width:62px; height:62px;
  display:flex; align-items:center; justify-content:center; border-radius:16px; }
.hero h1{ color:#fff !important; margin:0; font-size:27px; }
.hero p{ margin:3px 0 0; color:#D7E3F7; font-size:13.5px; }

/* Section header */
.sect{ display:flex; align-items:center; gap:11px; margin:6px 0 12px; }
.sect .badge{ background:var(--navy); color:#fff; width:28px; height:28px; border-radius:9px;
  display:flex; align-items:center; justify-content:center; font-family:'Sora'; font-weight:700; font-size:14px; }
.sect .t{ font-family:'Sora'; font-weight:600; font-size:18px; color:var(--navy); }

/* Kad seksyen */
[data-testid="stVerticalBlockBorderWrapper"]{ background:var(--card); border:1px solid var(--line) !important;
  border-radius:16px; box-shadow:0 4px 18px rgba(20,48,92,.06); }

/* Butang */
.stButton>button, .stDownloadButton>button{
  background:var(--navy); color:#fff; border:0; border-radius:11px; padding:.55rem 1.1rem;
  font-family:'Sora'; font-weight:600; font-size:14px; transition:.18s; box-shadow:0 4px 12px rgba(20,48,92,.18); }
.stButton>button:hover, .stDownloadButton>button:hover{ background:var(--navy2); transform:translateY(-1px);
  box-shadow:0 7px 18px rgba(20,48,92,.26); color:#fff; }
.stLinkButton>a{ background:linear-gradient(120deg,var(--navy),var(--navy2)); color:#fff !important;
  border-radius:11px; padding:.7rem 1.4rem; font-family:'Sora'; font-weight:600; text-decoration:none;
  box-shadow:0 6px 16px rgba(20,48,92,.22); transition:.18s; }
.stLinkButton>a:hover{ transform:translateY(-1px); box-shadow:0 9px 22px rgba(20,48,92,.3); }

/* Metrik */
[data-testid="stMetric"]{ background:var(--card); border:1px solid var(--line); border-radius:14px;
  padding:14px 16px; box-shadow:0 3px 12px rgba(20,48,92,.05); }
[data-testid="stMetricLabel"]{ color:var(--muted); font-weight:600; }
[data-testid="stMetricValue"]{ color:var(--navy); font-family:'Sora'; }

/* Status pill */
.pill{ display:inline-flex; align-items:center; gap:8px; background:#E8F5EE; color:#1B7A46;
  border:1px solid #BCE3CC; border-radius:999px; padding:6px 14px; font-size:13px; font-weight:600; }
.pill .dot{ width:9px; height:9px; border-radius:50%; background:#22A35A; box-shadow:0 0 0 3px rgba(34,163,90,.18); }

/* Uploader & dataframe */
[data-testid="stFileUploader"]{ background:#F7F9FD; border:1.5px dashed #C5D2E8; border-radius:14px; padding:8px; }
[data-testid="stDataFrame"]{ border-radius:12px; overflow:hidden; border:1px solid var(--line); }

/* Footer tersuai */
.foot{ text-align:center; color:var(--muted); font-size:12.5px; margin-top:26px;
  padding-top:16px; border-top:1px solid var(--line); }
.foot a{ color:var(--accent); text-decoration:none; font-weight:600; }
</style>
""", unsafe_allow_html=True)


def section(num, title):
    st.markdown(
        f'<div class="sect"><div class="badge">{num}</div><div class="t">{title}</div></div>',
        unsafe_allow_html=True,
    )


def hero():
    st.markdown(
        '<div class="hero"><div class="mark">🧾</div>'
        f'<div><h1>{APP_NAME} <span style="font-size:16px;opacity:.8">({APP_SHORT})</span></h1>'
        '<p>Resit Renyuk, Pudar atau Tulisan Tangan? Ambil Gambar Sahaja. '
        'Tak perlu key in satu-satu, AI Akan Memahami, Menyimpan dan '
        'Merekodkannya Secara Automatik.</p></div></div>',
        unsafe_allow_html=True,
    )


def footer():
    st.markdown(
        f'<div class="foot">\u00A9 {AUTHOR_YEAR} Dibangunkan oleh <b>{AUTHOR_NAME}</b> · '
        f'<a href="mailto:{AUTHOR_EMAIL}">{AUTHOR_EMAIL}</a><br>'
        f'<a href="{TOS_URL}" target="_blank">Terma Perkhidmatan</a> · '
        f'<a href="{PRIVACY_URL}" target="_blank">Dasar Privasi</a></div>',
        unsafe_allow_html=True,
    )


# ==================================================
# STATE + OAUTH REDIRECT
# ==================================================
for key, default in [("transactions", []), ("pending", [])]:
    if key not in st.session_state:
        st.session_state[key] = default

params = st.query_params
if "code" in params and "credentials" not in st.session_state:
    try:
        flow = get_flow()
        flow.fetch_token(code=params["code"])
        st.session_state.credentials = flow.credentials
        st.query_params.clear()
        st.rerun()
    except Exception as e:
        st.error(f"Gagal log masuk: {e}")

creds = st.session_state.get("credentials")
hero()

# ==================================================
# PINTU MASUK (log masuk wajib)
# ==================================================
if not creds:
    with st.container(border=True):
        st.markdown("#### 🔐 Log masuk diperlukan")
        st.write("Sila log masuk dengan akaun Google anda untuk mula menggunakan aplikasi. "
                 "Hanya pengguna yang dibenarkan boleh mengakses.")
        flow = get_flow()
        auth_url, _ = flow.authorization_url(
            prompt="consent", access_type="offline", include_granted_scopes="true"
        )
        st.link_button("Log masuk dengan Google", auth_url)
    footer()
    st.stop()

# ==================================================
# BAR STATUS + PAPAN PEMUKA METRIK
# ==================================================
service = drive_service(creds)

# ==================================================
# KAWALAN AKSES / TRIAL (berasaskan emel)
# ==================================================
if "user_email" not in st.session_state:
    st.session_state.user_email = get_user_email(creds)
user_email = st.session_state.user_email
access_status, days_left = check_access(user_email)

if access_status == "denied":
    with st.container(border=True):
        st.markdown("#### 🔒 Akaun belum didaftarkan")
        st.write(f"Emel **{user_email or 'anda'}** belum didaftarkan untuk ETS. "
                 "Sila hubungi kami untuk memulakan percubaan percuma.")
        st.markdown(f"📧 [{AUTHOR_EMAIL}](mailto:{AUTHOR_EMAIL})")
    footer()
    st.stop()

if access_status == "expired":
    with st.container(border=True):
        st.markdown("#### ⏰ Tempoh percubaan tamat")
        st.write("Terima kasih mencuba ETS! Untuk teruskan, sila langgan pelan bulanan atau tahunan.")
        st.markdown(f"📧 Hubungi untuk langgan: [{AUTHOR_EMAIL}](mailto:{AUTHOR_EMAIL})")
    footer()
    st.stop()

if access_status == "trial":
    st.info(f"🎁 Mod Percubaan — **{days_left} hari** lagi. "
            "Naik taraf ke pelan bulanan/tahunan untuk akses berterusan.")

# Muat profil pengguna sekali (dari Drive)
if "profile" not in st.session_state:
    st.session_state.profile = load_profile(service)
profile = st.session_state.profile

txns = st.session_state.transactions
total_amount = sum(float(t.get("amount", 0) or 0) for t in txns)
this_month = pd.Timestamp.now().strftime("%Y-%m")
month_amount = sum(
    float(t.get("amount", 0) or 0) for t in txns
    if str(t.get("transaction_date", "")).startswith(this_month)
)

c1, c2 = st.columns([3, 1])
with c1:
    who = profile.get("full_name") or "Pengguna"
    if profile.get("company"):
        who += f" · {profile['company']}"
    st.markdown(
        f'<span class="pill"><span class="dot"></span>{who}</span>',
        unsafe_allow_html=True)
with c2:
    if st.button("Log keluar"):
        st.session_state.clear()
        st.rerun()

# Ruangan Profil (boleh kemas kini)
with st.expander("👤 Profil Pengguna" + (" — sila isi" if not profile.get("full_name") else "")):
    pf_name = st.text_input("Nama Penuh", value=profile.get("full_name", ""))
    pf_company = st.text_input("Nama Syarikat (pilihan)", value=profile.get("company", ""))
    if st.button("💾 Simpan Profil"):
        new_profile = {"full_name": pf_name.strip(), "company": pf_company.strip()}
        st.session_state.profile = new_profile
        try:
            save_profile(service, new_profile)
            st.success("Profil disimpan.")
        except Exception as e:
            st.warning(f"Profil disimpan dalam sesi, tetapi gagal simpan ke Drive: {e}")
        st.rerun()

m1, m2, m3 = st.columns(3)
m1.metric("Jumlah Transaksi", f"{len(txns)}")
m2.metric("Jumlah Perbelanjaan", f"RM {total_amount:,.2f}")
m3.metric(f"Bulan {this_month}", f"RM {month_amount:,.2f}")

# Auto-backup: simpan automatik ke Drive setiap kali ada entri baru
st.session_state.auto_backup = st.toggle(
    "🔄 Auto-backup ke Drive selepas setiap entri baru",
    value=st.session_state.get("auto_backup", False),
    help="Bila aktif, data disimpan automatik ke fail backup bergulir di Drive — elak kehilangan data.",
)

st.write("")

# ==================================================
# SEKSYEN 1: MUAT FAIL INDUK
# ==================================================
with st.container(border=True):
    section("1", "Sambung Fail Induk")
    st.caption("Klik di awal hari untuk sambung rekod semalam. Memuat fail induk TERKINI dari Drive.")
    if st.button("📂 Muat fail induk terkini dari Drive"):
        with st.spinner("Mencari fail induk terkini di Drive..."):
            try:
                latest = find_latest_master(service)
                if latest:
                    st.session_state.transactions = read_master_bytes(
                        download_master(service, latest["id"]))
                    st.success(
                        f"{len(st.session_state.transactions)} transaksi dimuatkan dari: "
                        f"{latest['name']}")
                    st.rerun()
                else:
                    st.info("Tiada fail induk dijumpai di Drive. Akan dicipta bila anda simpan kali pertama.")
            except Exception as e:
                st.error(f"Gagal memuat fail induk: {e}")

# ==================================================
# SEKSYEN 2: UPLOAD RESIT
# ==================================================
with st.container(border=True):
    section("2", "Upload Resit Baru")
    st.caption("Boleh upload banyak gambar sekali gus. AI akan baca setiap satu.")
    receipts = st.file_uploader("Pilih gambar resit", type=["jpg", "jpeg", "png"],
                                accept_multiple_files=True, label_visibility="collapsed")
    if receipts and st.button("🔍 Analisa Semua Resit"):
        progress = st.progress(0)
        added, exact, pending = 0, 0, []
        arc_folder = get_or_create_folder(service, softcopy_folder_name())
        seq = count_archive_today(service, arc_folder)
        for i, file in enumerate(receipts):
            try:
                raw = file.getvalue()
                ext = (file.name.rsplit(".", 1)[-1] if "." in file.name else "jpg").lower()
                mime = file.type or "image/jpeg"
                data = extract_receipt_data(Image.open(BytesIO(raw)))
                status = classify(data, st.session_state.transactions)
                if status == "new":
                    st.session_state.transactions.append(data)
                    added += 1
                    seq += 1
                    save_archive_file(service, raw, arc_folder, seq, ext, mime)
                    st.write(f"✅ {file.name} — {data.get('vendor_name','?')} (RM{data.get('amount','?')})")
                elif status == "exact":
                    exact += 1
                    st.warning(f"⚠️ {file.name} — pendua tepat (no resit sama). Dilangkau.")
                else:
                    pending.append({"data": data, "name": file.name,
                                    "bytes": raw, "ext": ext, "mime": mime})
            except Exception as e:
                st.error(f"❌ {file.name}: {e}")
            progress.progress((i + 1) / len(receipts))
        st.session_state.pending = pending
        st.success(f"{added} ditambah · {exact} pendua dilangkau · {len(pending)} perlu disemak.")
        if added:
            maybe_autobackup(service)

# ==================================================
# SEKSYEN 3: TAMBAH TRANSAKSI MANUAL (TANPA RESIT)
# ==================================================
with st.container(border=True):
    section("3", "Tambah Transaksi Manual (Tanpa Resit)")
    st.caption("Untuk transaksi tanpa dokumen — contoh tol, parking, tip. Isi maklumat dan tambah.")

    # Kategori akaun debit yang seragam (chart of accounts ringkas)
    KATEGORI = [
        "Toll & Parking", "Travel Expense", "Fuel", "Office Supplies",
        "Utilities", "Meals & Entertainment", "Miscellaneous Expense", "Lain-lain",
    ]

    with st.form("manual_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            m_date = st.date_input("Tarikh")
            m_vendor = st.text_input("Vendor / Tempat", placeholder="cth: Tol PLUS, Parking DBKL")
            m_amount = st.number_input("Jumlah (RM)", min_value=0.0, step=0.50, format="%.2f")
        with c2:
            m_debit = st.selectbox("Akaun Debit (kategori)", KATEGORI)
            m_debit_lain = st.text_input("Jika 'Lain-lain', nyatakan kategori", placeholder="cth: Donation Expense")
            m_credit = st.text_input("Akaun Kredit", value="Cash")
        m_desc = st.text_input("Keterangan", placeholder="cth: Tol perjalanan ke pejabat klien")
        submitted = st.form_submit_button("➕ Tambah Transaksi")

    if submitted:
        if m_amount <= 0:
            st.error("Sila masukkan jumlah lebih daripada 0.")
        elif not m_vendor.strip():
            st.error("Sila isi Vendor / Tempat.")
        else:
            debit = m_debit_lain.strip() if (m_debit == "Lain-lain" and m_debit_lain.strip()) else m_debit
            serial = f"MV-{pd.Timestamp.now().strftime('%Y%m%d-%H%M%S')}"  # no siri unik
            txn = {
                "transaction_date": str(m_date),
                "vendor_name": m_vendor.strip(),
                "receipt_no": serial,  # no siri baucar manual
                "description": m_desc.strip() or m_vendor.strip(),
                "amount": float(m_amount),
                "debit_account": debit,
                "credit_account": m_credit.strip() or "Cash",
                "currency": "MYR",
            }
            st.session_state.transactions.append(txn)
            # Arkib baucar ke folder Resit Softcopy
            arc_folder = get_or_create_folder(service, softcopy_folder_name())
            seq = count_archive_today(service, arc_folder) + 1
            save_archive_file(service, make_voucher_image(txn), arc_folder, seq, "png", "image/png")
            maybe_autobackup(service)
            st.success(f"Ditambah: {txn['vendor_name']} — RM{txn['amount']:.2f} · No. Siri: {serial}")
            st.rerun()

# ==================================================
# SEMAKAN MANUSIA
# ==================================================
if st.session_state.pending:
    with st.container(border=True):
        section("!", "Perlu Disahkan")
        st.caption("Transaksi ini serupa & tiada nombor resit. Tanda yang BUKAN pendua.")
        for idx, item in enumerate(st.session_state.pending):
            d = item["data"]
            st.checkbox(
                f"{item['name']} — {d.get('vendor_name','?')} · RM{d.get('amount','?')} · {d.get('transaction_date','?')}",
                key=f"keep_{idx}",
            )
        if st.button("✅ Sahkan pilihan"):
            kept = 0
            arc_folder = get_or_create_folder(service, softcopy_folder_name())
            seq = count_archive_today(service, arc_folder)
            for idx, item in enumerate(st.session_state.pending):
                if st.session_state.get(f"keep_{idx}"):
                    st.session_state.transactions.append(item["data"])
                    seq += 1
                    save_archive_file(service, item["bytes"], arc_folder, seq,
                                      item.get("ext", "jpg"), item.get("mime", "image/jpeg"))
                    kept += 1
            st.session_state.pending = []
            st.success(f"{kept} transaksi ditambah sebagai pembelian berasingan.")
            if kept:
                maybe_autobackup(service)
            st.rerun()

# ==================================================
# SEKSYEN 4: FAIL INDUK & LAPORAN
# ==================================================
with st.container(border=True):
    section("4", "Fail Induk & Laporan")
    if st.session_state.transactions:
        st.dataframe(pd.DataFrame(st.session_state.transactions), use_container_width=True)
        excel_bytes = build_master_excel(
            st.session_state.transactions, st.session_state.get("profile")).getvalue()
        folder_name = monthly_folder_name()
        st.caption(f"Akan disimpan ke folder Drive: **{folder_name}**")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("☁️ Simpan ke Google Drive"):
                folder_id = get_or_create_folder(service, folder_name)
                filename = next_filename(service, folder_id)
                created = save_master_in_folder(service, excel_bytes, folder_id, filename)
                st.success(f"Disimpan: {created['name']}\n\n→ folder: {folder_name} ✅")
        with col2:
            dl_name = f"{MASTER_PREFIX}_{pd.Timestamp.now().strftime('%d%m%Y')}.xlsx"
            st.download_button("📥 Muat turun (sandaran)", data=excel_bytes,
                               file_name=dl_name, mime=XLSX_MIME,
                               use_container_width=True)
    else:
        st.info("Belum ada transaksi. Muat fail induk atau upload resit di atas.")

footer()
